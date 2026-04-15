#!/usr/bin/env python3
"""Aggregate CAPES artpe authors per article for INPE and IN_GLOSA=0."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook


FILE_PATTERN = "br-capes-colsucup-prod-autor-*-bibliografica-artpe-*.xlsx"
REQUIRED_COLUMNS = [
    "AN_BASE",
    "CD_PROGRAMA_IES",
    "NM_PROGRAMA_IES",
    "SG_ENTIDADE_ENSINO",
    "ID_ADD_PRODUCAO_INTELECTUAL",
    "TP_AUTOR",
    "IN_GLOSA",
]
ESCAPE_RE = re.compile(r"_x([0-9A-Fa-f]{4})_")


def decode_excel_escapes(value: str) -> str:
    text = str(value)
    previous = None
    while previous != text:
        previous = text
        text = ESCAPE_RE.sub(lambda m: chr(int(m.group(1), 16)), text)
    return text


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return decode_excel_escapes(str(value)).strip()


def normalize_year(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    try:
        number = float(text.replace(",", "."))
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass
    match = re.search(r"(19|20)\d{2}", text)
    if match:
        return match.group(0)
    return text


def normalize_integral_id(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    try:
        number = float(text.replace(",", "."))
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass
    return text


def is_glosa_zero(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip().replace(",", ".")
    if text == "":
        return False
    try:
        return float(text) == 0.0
    except ValueError:
        return text == "0"


def to_ascii_slug(text: str) -> str:
    folded = unicodedata.normalize("NFD", text)
    folded = "".join(ch for ch in folded if unicodedata.category(ch) != "Mn")
    slug = "".join(ch.lower() if ch.isalnum() else "_" for ch in folded.strip())
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_")


def discover_files(raw_dir: Path) -> List[Path]:
    return sorted(p for p in raw_dir.glob(FILE_PATTERN) if not p.name.startswith("~$"))


def discover_files_with_pattern(raw_dir: Path, file_pattern: str) -> List[Path]:
    return sorted(p for p in raw_dir.glob(file_pattern) if not p.name.startswith("~$"))


def read_header_map(sheet) -> Dict[str, int]:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping: Dict[str, int] = {}
    for idx, col in enumerate(header):
        if col is None:
            continue
        normalized = decode_excel_escapes(str(col)).strip()
        if normalized and normalized not in mapping:
            mapping[normalized] = idx
    return mapping


def generate_column_map(tp_values: Iterable[str]) -> Dict[str, str]:
    used: set[str] = set()
    col_map: Dict[str, str] = {}
    for tp in sorted(tp_values):
        base = to_ascii_slug(tp) or "sem_informacao"
        candidate = f"qtd_tp_autor_{base}"
        if candidate in used:
            i = 2
            while f"{candidate}_{i}" in used:
                i += 1
            candidate = f"{candidate}_{i}"
        used.add(candidate)
        col_map[tp] = candidate
    return col_map


def safe_int(text: str) -> int:
    try:
        return int(float(text))
    except ValueError:
        return 999999


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Aggregate CAPES artpe production-author rows into one line per article "
            "with TP_AUTOR counts, filtered by SG_ENTIDADE_ENSINO and IN_GLOSA=0."
        )
    )
    parser.add_argument(
        "--raw-dir",
        type=Path,
        default=Path("RawData"),
        help="Directory containing CAPES xlsx files.",
    )
    parser.add_argument(
        "--file-pattern",
        default=FILE_PATTERN,
        help=f"Glob pattern for input files. Default: {FILE_PATTERN}",
    )
    parser.add_argument(
        "--out-csv",
        type=Path,
        default=Path("Data/inpe_artpe_autores_por_artigo_2013_2024_glosa0.csv"),
        help="Output CSV path (one record per article).",
    )
    parser.add_argument(
        "--out-tp-csv",
        type=Path,
        default=Path("Data/inpe_artpe_tp_autor_unicos_2013_2024_glosa0.csv"),
        help="Output CSV path with unique TP_AUTOR values and mapping.",
    )
    parser.add_argument(
        "--out-log-json",
        type=Path,
        default=Path("Data/logs/inpe_artpe_autores_por_artigo_2013_2024_glosa0_log.json"),
        help="Output JSON log path.",
    )
    parser.add_argument(
        "--target-sg",
        default="INPE",
        help="Target SG_ENTIDADE_ENSINO value. Default: INPE.",
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=250000,
        help="Print progress every N scanned rows per file.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    raw_dir: Path = args.raw_dir
    target_sg = args.target_sg.strip().upper()
    file_pattern = str(args.file_pattern).strip() or FILE_PATTERN

    files = discover_files_with_pattern(raw_dir, file_pattern)
    if not files:
        raise SystemExit(f"No files found in {raw_dir} matching {file_pattern}.")

    article_meta: Dict[str, Dict[str, str]] = {}
    article_tp_counts: Dict[str, Counter] = defaultdict(Counter)
    tp_values: set[str] = set()

    tp_total_rows = Counter()
    tp_total_articles = Counter()

    per_file_stats = []
    total_scanned = 0
    total_nonempty = 0
    total_filtered = 0
    total_skipped_missing_id = 0
    meta_conflicts = 0
    years_present = set()

    print(f"Found {len(files)} files.")
    for file_path in files:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_map = read_header_map(ws)
        missing = [col for col in REQUIRED_COLUMNS if col not in header_map]
        if missing:
            wb.close()
            raise SystemExit(f"Missing columns in {file_path.name}: {missing}")

        idx = header_map
        scanned = 0
        nonempty = 0
        filtered = 0
        skipped_missing_id = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            scanned += 1

            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            nonempty += 1

            sg = normalize_text(row[idx["SG_ENTIDADE_ENSINO"]]).upper()
            if sg != target_sg:
                continue
            if not is_glosa_zero(row[idx["IN_GLOSA"]]):
                continue

            article_id = normalize_integral_id(row[idx["ID_ADD_PRODUCAO_INTELECTUAL"]])
            if not article_id:
                skipped_missing_id += 1
                continue

            year = normalize_year(row[idx["AN_BASE"]])
            program_code = normalize_text(row[idx["CD_PROGRAMA_IES"]])
            program_name = normalize_text(row[idx["NM_PROGRAMA_IES"]])
            tp_autor = normalize_text(row[idx["TP_AUTOR"]]) or "SEM_INFORMACAO"

            if year:
                years_present.add(year)
            tp_values.add(tp_autor)
            tp_total_rows[tp_autor] += 1

            if article_id not in article_meta:
                article_meta[article_id] = {
                    "an_base": year,
                    "sg_entidade_ensino": sg,
                    "cd_programa_ies": program_code,
                    "nm_programa_ies": program_name,
                }
            else:
                meta = article_meta[article_id]
                if (
                    (year and meta["an_base"] and year != meta["an_base"])
                    or (
                        program_code
                        and meta["cd_programa_ies"]
                        and program_code != meta["cd_programa_ies"]
                    )
                    or (
                        program_name
                        and meta["nm_programa_ies"]
                        and program_name != meta["nm_programa_ies"]
                    )
                ):
                    meta_conflicts += 1

            article_tp_counts[article_id][tp_autor] += 1
            filtered += 1

            if args.progress_every > 0 and scanned % args.progress_every == 0:
                print(
                    f"[{file_path.name}] scanned={scanned} nonempty={nonempty} "
                    f"filtered={filtered} artigos={len(article_meta)}"
                )

        wb.close()
        total_scanned += scanned
        total_nonempty += nonempty
        total_filtered += filtered
        total_skipped_missing_id += skipped_missing_id

        file_stats = {
            "file": file_path.name,
            "rows_scanned": scanned,
            "rows_nonempty": nonempty,
            "rows_filtered_target_glosa0": filtered,
            "rows_skipped_missing_article_id": skipped_missing_id,
        }
        per_file_stats.append(file_stats)
        print(
            f"[done] {file_path.name}: scanned={scanned} nonempty={nonempty} "
            f"filtered={filtered} artigos_unicos_atual={len(article_meta)}"
        )

    column_map = generate_column_map(tp_values)

    for article_id, counts in article_tp_counts.items():
        for tp in counts:
            if counts[tp] > 0:
                tp_total_articles[tp] += 1

    base_columns = [
        "id_add_producao_intelectual",
        "an_base",
        "sg_entidade_ensino",
        "cd_programa_ies",
        "nm_programa_ies",
        "qtd_autores_total",
    ]
    count_columns = [column_map[tp] for tp in sorted(tp_values)]

    output_rows = []
    for article_id, meta in article_meta.items():
        counts = article_tp_counts[article_id]
        row = {
            "id_add_producao_intelectual": article_id,
            "an_base": meta["an_base"],
            "sg_entidade_ensino": meta["sg_entidade_ensino"],
            "cd_programa_ies": meta["cd_programa_ies"],
            "nm_programa_ies": meta["nm_programa_ies"],
            "qtd_autores_total": sum(counts.values()),
        }
        for tp in sorted(tp_values):
            row[column_map[tp]] = counts.get(tp, 0)
        output_rows.append(row)

    output_rows.sort(
        key=lambda r: (safe_int(str(r["an_base"])), str(r["id_add_producao_intelectual"]))
    )

    args.out_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=base_columns + count_columns)
        writer.writeheader()
        writer.writerows(output_rows)

    args.out_tp_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.out_tp_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "tp_autor",
                "coluna_csv",
                "qtd_linhas_autor",
                "qtd_artigos_com_categoria",
            ],
        )
        writer.writeheader()
        for tp in sorted(tp_values):
            writer.writerow(
                {
                    "tp_autor": tp,
                    "coluna_csv": column_map[tp],
                    "qtd_linhas_autor": tp_total_rows.get(tp, 0),
                    "qtd_artigos_com_categoria": tp_total_articles.get(tp, 0),
                }
            )

    years_sorted = sorted(years_present, key=safe_int)
    log_payload = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "target_sg_entidade_ensino": target_sg,
        "filter_in_glosa": 0,
        "input_file_pattern": file_pattern,
        "input_file_count": len(files),
        "input_files": [p.name for p in files],
        "rows_scanned_total": total_scanned,
        "rows_nonempty_total": total_nonempty,
        "rows_filtered_target_glosa0_total": total_filtered,
        "rows_skipped_missing_article_id_total": total_skipped_missing_id,
        "unique_articles": len(article_meta),
        "unique_tp_autor_values": sorted(tp_values),
        "years_present": years_sorted,
        "meta_conflicts": meta_conflicts,
        "output_csv": str(args.out_csv),
        "output_tp_csv": str(args.out_tp_csv),
        "per_file_stats": per_file_stats,
    }
    args.out_log_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_log_json.write_text(
        json.dumps(log_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print("\nExtraction completed.")
    print(f"Output articles: {args.out_csv}")
    print(f"Output TP_AUTOR: {args.out_tp_csv}")
    print(f"Output log: {args.out_log_json}")
    print(f"Unique articles: {len(article_meta)}")
    print(f"Years covered: {years_sorted[0]} to {years_sorted[-1]}" if years_sorted else "No years.")
    print(f"TP_AUTOR values: {sorted(tp_values)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
